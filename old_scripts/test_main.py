# this is the main script, the idea is to invoke the dependencies queries, functions and headers from other scripts

# import libraries

import requests
import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

# invoke dependencies
import function_query
from function_query import query_pdc # function to retrieve data
import test_queries # import queries list and subfunctions
from test_queries import *
import test_headers
from test_headers import *
import test_for_loops
from test_for_loops import *

# variables added in PDC portal

try:
    pdc_study_id_input = str(input("Introduce a PDC study id, ie: PDC000443: "))
    #offset_input = int(input("introduce offset, ie: 0 "))
    #data_type_input = str(input("data_type, try: log2_ratio: "))
except ValueError:
    print("Please enter a valid id.")


#pdc_study_id_input = "PDC000219"
#offset_input = 0 # set offset records to pull
#limit_input = 10 # limit number of records,
#data_type_input = 'log2_ratio' # option: unshared_log2_ratio 

# variables inhereted

# Variables
variables = {
    "pdc_study_identifier": pdc_study_id_input,
    "offset": 0, 
    "limit": 20750,  
   #"data_type": data_type_input 
}

# Project-Program 
study_data = query_pdc(query= query_study_info, variables=variables)
matrix = json.loads(study_data.content)['data']['study']
study_df = pd.DataFrame(matrix)
program_project = study_df[program_project_header].transpose()
program_project.columns = ['name']
program_project[" "] = program_project.index
program_project = program_project.reindex(columns=[' ', 'name'])


# Case-Matrix
speciment_data = query_pdc(query= query_biospecimen, variables=variables)
matrix = json.loads(speciment_data.content)['data']["biospecimenPerStudy"]
biospecimen_df = pd.DataFrame(matrix[1:], columns=matrix[0])
case_matrix = biospecimen_df[case_matrix_header]
case_matrix = case_matrix[2:]


# Case
case_data = query_pdc(query= query_case, variables=variables)
matrix = json.loads(case_data.content)['data']['case']
case_data_df = pd.DataFrame(matrix)
case = pd.merge(left=biospecimen_df, right=case_data_df, on="case_id")
columns_to_keep = [col for col in case.columns if not col.endswith('_y')]
case = case[columns_to_keep]
case = case.rename(columns={col: col.rstrip('_x') for col in case.columns})
to_remove = list(set(case.columns).difference(case_header))
case.drop(columns=to_remove, inplace=True)
case = case.reindex(columns=case_header)

# Demographic

variables['study_id'] = study_df['study_id'][0]
demographics_data = query_pdc(query= query_demographcis, variables=variables)
matrix = json.loads(demographics_data.content)['data']["paginatedCaseDemographicsPerStudy"]["caseDemographicsPerStudy"]
demographics_data = pd.DataFrame(matrix[1:], columns=matrix[0])
demographics_data['demographic_id'] = demographics_data['demographics'].apply(lambda diag_list: diag_list[0]['demographic_id'] if diag_list else None)
demographics_df = for_demographics(matrix = matrix)
demographic = pd.merge(left=demographics_data, right=demographics_df, on="demographic_id")
to_remove = list(set(demographic.columns).difference(demographics_header))
demographic.drop(columns=to_remove, inplace=True)
demographic = demographic.reindex(columns=demographics_header)

# Diganosis
diagnose_data = query_pdc(query= query_diagnose, variables=variables)
matrix = json.loads(diagnose_data.content)['data']["paginatedCaseDiagnosesPerStudy"]["caseDiagnosesPerStudy"]
diagnose_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
diagnose_data_df['diagnosis_id'] = diagnose_data_df['diagnoses'].apply(lambda diag_list: diag_list[0]['diagnosis_id'] if diag_list else None)
diagnose_df = for_diagnosis(matrix = matrix)
diagnosis = pd.merge(left=diagnose_data_df, right=diagnose_df, on="diagnosis_id")
to_remove = list(set(diagnosis.columns).difference(diagnose_header))
diagnosis.drop(columns=to_remove, inplace=True)
diagnosis = diagnosis.reindex(columns=diagnose_header)

# Exposure

exposure_data = query_pdc(query= query_exposure, variables= variables)
matrix = json.loads(exposure_data.content)['data']["paginatedCaseExposuresPerStudy"]["caseExposuresPerStudy"]
exposure_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
exposure_data_df['exposure_id'] = exposure_data_df['exposures'].apply(lambda diag_list: diag_list[0]['exposure_id'] if diag_list else None)
exposure_df = for_exposure(matrix = matrix)
exposure = pd.merge(left=exposure_data_df, right=exposure_df, on="exposure_id")
to_remove = list(set(exposure.columns).difference(expose_header))
exposure = exposure.drop(columns=to_remove)
exposure = exposure.reindex(columns=expose_header)

# Treatment

treatments_data = query_pdc(query= query_treatments, variables=variables)
matrix = json.loads(treatments_data.content)['data']["paginatedCaseTreatmentsPerStudy"]["caseTreatmentsPerStudy"]
treatments_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
treatments_data_df['treatment_id'] = treatments_data_df['treatments'].apply(lambda diag_list: diag_list[0]['treatment_id'] if diag_list else None)
treatments_df = for_treatment(matrix = matrix)
treatments = pd.merge(left=treatments_data_df, right=treatments_df, on="treatment_id")
to_remove = list(set(treatments.columns).difference(treatment_header))
treatments = treatments.drop(columns=to_remove)
treatments = treatments.reindex(columns=treatment_header)

# Follow up

follow_up_data = query_pdc(query= query_follow_up, variables=variables)
matrix = json.loads(follow_up_data.content)['data']["paginatedCaseFollowUpsPerStudy"]['caseFollowUpsPerStudy']
follow_up_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
follow_up_data_df['follow_up_id'] = follow_up_data_df['follow_ups'].apply(lambda diag_list: diag_list[0]['follow_up_id'] if diag_list else None)
follow_up_df = for_follows_up(matrix = matrix)
follow_ups = pd.merge(left=follow_up_data_df, right=follow_up_df, on="follow_up_id")
to_remove = list(set(follow_ups.columns).difference(follow_ups_header))
follow_ups = follow_ups.drop(columns=to_remove)
follow_ups = follow_ups.reindex(columns=follow_ups_header)

# Sample

matrix = json.loads(case_data.content)['data']['case']
case_data_df = pd.DataFrame(matrix)
sample_df = for_sample(matrix = matrix)
to_remove = list(set(sample_df.columns).difference(sample_header))
samples = sample_df.drop(columns=to_remove)
samples = samples.reindex(columns=sample_header)


# Aliquots
aliquots_data = query_pdc(query= query_aliquots, variables=variables)
matrix = json.loads(aliquots_data.content)["data"]["paginatedCasesSamplesAliquots"]["casesSamplesAliquots"]
aliquots_df = pd.DataFrame(matrix)
aliquots_df['gdc_sample_id'] = aliquots_df['samples'].apply(lambda diag_list: diag_list[0]['gdc_sample_id'] if diag_list else None)
tmp = for_aliquots(matrix= matrix)
aliquots_df = pd.merge(left=biospecimen_df, right=tmp, on = "aliquot_submitter_id",suffixes= ("", "_"))
to_remove = list(set(aliquots_df.columns).difference(aliquots_header))
aliquots_df = aliquots_df.drop(columns=to_remove)
aliquots = aliquots_df.reindex(columns=aliquots_header)

# Study - Data
matrix = json.loads(study_data.content)['data']['study']
study_df = pd.DataFrame(matrix)
to_remove = list(set(study_df.columns).difference(study_header))
study = study_df.drop(columns=to_remove)
study = study.reindex(columns=study_header)

# Protocol

protocol_Data = query_pdc(query= query_protocol, variables=variables)
matrix = json.loads(protocol_Data.content)['data']['protocolPerStudy']
protocol_df = pd.DataFrame(matrix)
to_remove = list(set(protocol_df.columns).difference(protocol_header))
protocol = protocol_df.drop(columns=to_remove)
protocol = protocol.reindex(columns=protocol_header)

# Experimental - Metadata

expMetadat_data_2 = query_pdc(query= query_expMetadata_2, variables=variables)
matrix = json.loads(expMetadat_data_2.content)['data']["studyExperimentalDesign"]
expMetadat_data_2 = pd.DataFrame(matrix)
to_remove = list(set(expMetadat_data_2.columns).difference(exp_metadata_header))
Exp_Metadata = expMetadat_data_2.drop(columns=to_remove)
Exp_Metadata = Exp_Metadata.reindex(columns=exp_metadata_header)
tmt_columns = [
    'tmt_126', 'tmt_127n', 'tmt_127c', 'tmt_128n', 'tmt_128c',
    'tmt_129n', 'tmt_129c', 'tmt_130n', 'tmt_130c', 'tmt_131', 'tmt_131c'
]

# Function to extract aliquot_ids
def extract_aliquots(cell):
    if isinstance(cell, list):
        return [entry.get('aliquot_id') for entry in cell if isinstance(entry, dict)]
    return []

# Function to format aliquot_ids
def format_aliquot_ids(aliquot_list):
    return [f"aliquot_id: {aliquot}" for aliquot in aliquot_list]

# Modify Exp_Metadata in-place
for col in tmt_columns:
    if col in Exp_Metadata.columns:
        Exp_Metadata[col] = Exp_Metadata[col].apply(extract_aliquots).map(format_aliquot_ids)



# File Metadata

file_metadata_data = query_pdc(query= query_file_metadata, variables=variables)
matrix = json.loads(file_metadata_data.content)['data']["filesPerStudy"]
file_metadata_df = pd.DataFrame(matrix)
to_remove = list(set(file_metadata_df.columns).difference(file_metadata_header))
file_metada = file_metadata_df.drop(columns=to_remove)
file_metada = file_metada.reindex(columns=file_metadata_header)


file_id_list = file_metadata_df["file_id"].dropna().astype(str).tolist()

test_files = file_id_list[4:10]

all_metadata = []

# Loop through each file_id
for file_id in file_id_list:
    variables_2 = {
        "file_id": file_id,
        "offset": 0,
        "limit": 20750
    }
    try:
        result = query_pdc(query=query_file_metadata_2, variables=variables_2)
        matrix = json.loads(result.content)['data']["fileMetadata"]
        all_metadata.extend(matrix)
    except Exception as e:
        print(f"Failed for file_id {file_id}: {e}")

all_metadata_df = pd.DataFrame(all_metadata)

tmp = pd.merge(file_metadata_df, all_metadata_df, on="file_id")
columns_to_keep = [col for col in tmp.columns if not col.endswith('_y')]
tmp = tmp[columns_to_keep]
tmp = tmp.rename(columns={col: col.rstrip('_x') for col in tmp.columns})
tmp = tmp.rename(columns={'plex_or_dataset_name': "plex_or_folder_name", 
                    "study_run_metadata_id": "run_metadata_id",
                    "protocol_submitter_id": "protocol"})
to_remove = list(set(tmp.columns).difference(file_metadata_header_2))
file_metada = tmp.drop(columns=to_remove)
file_metada = file_metada.reindex(columns=file_metadata_header_2)



# File Metadata_2
expMetadat_data_2.columns
file_metadata_df.columns
file_metadata_df["run_metadata_id"]
file_metadata_header_2
expMetadat_data_2.rename(columns={"plex_dataset_name": "plex_or_folder_name",
                                  "study_run_metadata_id": "run_metadata_id",
                                  "protocol_submitter_id": "protocol"})
expMetadat_data_2["plex_dataset_name"]
Exp_Metadata.columns

# object dictionary:
study_information = {
    "Project-Program": pd.DataFrame(program_project),
    "Case-Matrix": pd.DataFrame(case_matrix),
    "Case": pd.DataFrame(case),
    "Demographic": pd.DataFrame(demographic),
    "Diagnosis": pd.DataFrame(diagnosis),
    "Exposure": pd.DataFrame(exposure),
    "Family-History": pd.DataFrame(), #need to find the data
    "Treatments": pd.DataFrame(treatments),
    "Follow-up": pd.DataFrame(follow_ups),
    "Sample": pd.DataFrame(samples), #need to redo
    "Aliquots": pd.DataFrame(aliquots), #need to redo
    "Study": pd.DataFrame(study),
    "Protocol": pd.DataFrame(protocol),
    "Exp_Metadata": pd.DataFrame(Exp_Metadata),
    "File-Metadata": pd.DataFrame(file_metada),
}

filename = f"Study_Data_{pdc_study_id_input}.xlsx"

with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
    for sheet_name, df in study_information.items():
        if df is not None and not df.empty:  # Check if the DataFrame is not None and not empty
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Create an empty DataFrame and write it to the sheet
            empty_df = pd.DataFrame(columns= df.columns, index=range(10)).fillna('data not available')
            empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create a study_summary page 
study_transposed = study_df.transpose()
study_transposed = study_transposed[:-1]

files_count_df = pd.DataFrame(study_df.loc[0, "filesCount"])

# Define a unique new sheet name
sheet_name = "Study Summary"

# Open writer to append a sheet (no custom 'book' logic)
with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    # Write the transposed study_df
    study_transposed.to_excel(writer, index=True, header=True, sheet_name=sheet_name, startrow=0)
    
    # Leave a gap and write the transposed filesCount
    start_row = study_transposed.shape[0] + 5
    files_count_df.to_excel(writer, index=False, header=True, sheet_name=sheet_name, startrow=start_row)

wb = load_workbook(filename)

sheet_order = wb.sheetnames
sheet_order.insert(0, sheet_order.pop(sheet_order.index("Study Summary")))
wb._sheets = [wb[sheet] for sheet in sheet_order]

wb.save(filename)



print('download completed')